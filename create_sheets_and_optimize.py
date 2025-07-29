# importing all requirements

import pandas as pd
import numpy as np
import xlwings as xw
import openpyxl as xl
from scipy.optimize import minimize
import math
import shutil
import os
import warnings
warnings.filterwarnings('ignore')

# functions

def excel_column_name(column_number):
    column_name = ""
    while column_number > 0:
        remainder = (column_number - 1) % 26
        column_name = chr(65 + remainder) + column_name
        column_number = (column_number - 1) // 26
    return column_name

def add_sheets(year, input_file_path, optimization=None, max_sha_weights=None, min_var_weights=None, max_ret_weights=None, expected_ann_returns=None, maximum_ann_stdev=None):
    """
    optimization can either be "MAX SHA" or "MIN VAR" or "MAX RET" or None
    """

    file_name = os.path.basename(input_file_path)
    file_path = input_file_path
    xl = pd.ExcelFile(file_path)

    df_weekly_data = xl.parse(0)
    df_weekly_returns = xl.parse(1)
    df_cov_matrix = xl.parse(2)

    df_weekly_returns_no_date = df_weekly_returns.iloc[:, 1:]

    average_returns = df_weekly_returns_no_date.mean()
    std_dev_returns = df_weekly_returns_no_date.std()

    df_summary_stats = pd.DataFrame([average_returns, std_dev_returns], index=['Average', 'Std Dev'])

    num_symbs = len(df_weekly_data.columns) - 1
    data_rows = len(df_weekly_returns)

    symbols = df_weekly_data.columns[1:].tolist()

    if optimization == None:
        df_max_sharpe_ratio = pd.DataFrame({'Symbol': symbols})
        df_max_sharpe_ratio['Weights'] = max_sha_weights

        df_min_variance = pd.DataFrame({'Symbol': symbols})
        df_min_variance['Weights'] = min_var_weights

        df_max_returns = pd.DataFrame({'Symbol': symbols})
        df_max_returns['Weights'] = max_ret_weights
    else:
        df_max_sharpe_ratio = xl.parse(4)
        df_min_variance = xl.parse(5)
        df_max_returns = xl.parse(6)

    output_file_path = os.path.dirname(input_file_path)+'/updated_'+file_name

    with pd.ExcelWriter(output_file_path, engine='xlsxwriter') as writer:

        df_weekly_data.to_excel(writer, sheet_name='weekly_historical_data', index=False)
        df_weekly_returns.to_excel(writer, sheet_name='weekly_returns', index=False)
        df_cov_matrix.to_excel(writer, sheet_name='covariance_matrix', index=False)
        df_summary_stats.to_excel(writer, sheet_name='summary_stats')
        df_max_sharpe_ratio.to_excel(writer, sheet_name='max_sharpe_ratio', index=False)
        df_min_variance.to_excel(writer, sheet_name='min_variance', index=False)
        df_max_returns.to_excel(writer, sheet_name='max_returns', index=False)

        summary_stats = writer.sheets['summary_stats']
        for col in range(1, num_symbs + 1):
            summary_stats.write(0, col, df_weekly_data.columns[col])
            summary_stats.write_formula(1, col, f"=AVERAGE('weekly_returns'!{excel_column_name(col+1)}2:{excel_column_name(col+1)}{data_rows+1})")

        # for max sharpe ratio
        
        worksheet = writer.sheets['max_sharpe_ratio']
            
        if optimization == None:
            for row_num in range(1, num_symbs+1):
                worksheet.write_formula(row_num, 1, f'{1/num_symbs}')
        else:
            for row_num in range(1, len(max_sha_weights)+1):
                worksheet.write_formula(row_num, 1, f'{max_sha_weights[row_num-1]}')
        
        worksheet.write('D2', '')
        worksheet.write('D3', 'Weekly Portfolio Return')
        worksheet.write('D4', 'Annualized Portfolio Return')
        worksheet.write('D5', 'Portfolio Variance')
        worksheet.write('D6', 'Portfolio Standard Deviation')
        worksheet.write('D8', 'Sharpe Ratio - Max')
        worksheet.write('D9', 'Sum of Weights')
        worksheet.write('D10', 'Risk Free Rate (Daily)')

        sum_of_weights_formula = f'=SUM(B2:B{num_symbs+1})'
        weekly_portfolio_return_formula = '{'+f"=SUMPRODUCT(TRANSPOSE(B2:B{num_symbs+1}),'summary_stats'!B2:{excel_column_name(num_symbs+1)}2)"+'}'
        portfolio_variance_formula = '{'+f"=MMULT(MMULT(TRANSPOSE(B2:B{num_symbs+1}),'covariance_matrix'!B2:{excel_column_name(num_symbs+1)}{num_symbs+1}),B2:B{num_symbs+1})"+'}'

        worksheet.write_formula('E3', weekly_portfolio_return_formula)
        worksheet.write_formula('E4', f'=((E3+1)^52)-1')
        worksheet.write_formula('E5', portfolio_variance_formula)
        worksheet.write_formula('E6', f'=SQRT(E5)')
        worksheet.write_formula('E8', f'=E3/E6')
        worksheet.write_formula('E9', sum_of_weights_formula)
        worksheet.write_formula('E10', str(nominal_risk_free_rate[year]))

        # for min variance

        worksheet = writer.sheets['min_variance']
        
        if optimization == None:
            for row_num in range(1, num_symbs+1):
                worksheet.write_formula(row_num, 1, f'{1/num_symbs}')
        elif optimization == 'MAX RET':
            for row_num in range(1, len(min_var_weights)+1):
                worksheet.write_formula(row_num, 1, f'{min_var_weights[row_num-1]}')

        worksheet.write('D2', '')
        worksheet.write('D3', 'Weekly Portfolio Return')
        worksheet.write('D4', 'Annualized Portfolio Return')
        worksheet.write('D5', 'Portfolio Variance')
        worksheet.write('D6', 'Portfolio Standard Deviation')
        worksheet.write('D8', 'Min Return Expectation')
        worksheet.write('D9', 'Sum of Weights')
        
        sum_of_weights_formula = f'=SUM(B2:B{num_symbs+1})'
        portfolio_return_formula = '{'+f"=SUMPRODUCT(TRANSPOSE(B2:B{num_symbs+1}),'summary_stats'!B2:{excel_column_name(num_symbs+1)}2)"+'}'
        portfolio_variance_formula = '{'+f"=MMULT(MMULT(TRANSPOSE(B2:B{num_symbs+1}),'covariance_matrix'!B2:{excel_column_name(num_symbs+1)}{num_symbs+1}),B2:B{num_symbs+1})"+'}'

        worksheet.write_formula('E3', portfolio_return_formula)
        worksheet.write_formula('E4', f'=((E3+1)^52)-1')
        worksheet.write_formula('E5', portfolio_variance_formula)
        worksheet.write_formula('E6', f'=SQRT(E5)')
        worksheet.write_formula('E8', f"={expected_ann_returns}")
        worksheet.write_formula('E9', sum_of_weights_formula)

        # for max returns

        worksheet = writer.sheets['max_returns']
        if optimization == 'MAX RET':
            for row_num in range(1, len(max_ret_weights)+1):
                worksheet.write_formula(row_num, 1, f'{max_ret_weights[row_num-1]}')
        elif optimization == None:
            for row_num in range(1, num_symbs+1):
                worksheet.write_formula(row_num, 1, f'{1/num_symbs}')
        
        worksheet.write('D2', '')
        worksheet.write('D3', 'Weekly Portfolio Return')
        worksheet.write('D4', 'Annualized Portfolio Return')
        worksheet.write('D5', 'Portfolio Variance')
        worksheet.write('D6', 'Portfolio Standard Deviation')
        worksheet.write('D8', 'Min Return Expectation')
        worksheet.write('D9', 'Sum of Weights')
        
        sum_of_weights_formula = f'=SUM(B2:B{num_symbs+1})'
        portfolio_return_formula = '{'+f"=SUMPRODUCT(TRANSPOSE(B2:B{num_symbs+1}),'summary_stats'!B2:{excel_column_name(num_symbs+1)}2)"+'}'
        portfolio_variance_formula = '{'+f"=MMULT(MMULT(TRANSPOSE(B2:B{num_symbs+1}),'covariance_matrix'!B2:{excel_column_name(num_symbs+1)}{num_symbs+1}),B2:B{num_symbs+1})"+'}'

        worksheet.write_formula('E3', portfolio_return_formula)
        worksheet.write_formula('E4', f'=((E3+1)^52)-1')
        worksheet.write_formula('E5', portfolio_variance_formula)
        worksheet.write_formula('E6', f'=SQRT(E5)')
        worksheet.write_formula('E8', f"={maximum_ann_stdev}")
        worksheet.write_formula('E9', sum_of_weights_formula)
    
    os.chmod(output_file_path, 0o666)

# Max Sharpe Ratio

def max_sharpe_ratio(average, cov_matrix, num_symbs, risk_free_rate):

    def calculate_e3(weights, average):
        return np.dot(weights, average)

    def calculate_e5(weights, cov_matrix):
        return np.dot(np.dot(weights, cov_matrix), weights)

    def calculate_e6(e5_value):
        return np.sqrt(e5_value)

    def calculate_e8(e3_value, e6_value):
        return (e3_value-risk_free_rate)/e6_value

    def objective_function(weights):
        e3 = calculate_e3(weights, average)
        e5 = calculate_e5(weights, cov_matrix)
        e6 = calculate_e6(e5)
        e8 = calculate_e8(e3, e6)
        return -e8

    weights = [1/num_symbs]*num_symbs

    # Constraints and bounds remain the same as before
    constraints = [{'type': 'eq', 'fun': lambda x: np.sum(x) - 1}]
    bounds = [(0, 1)] * len(weights)

    # Perform the optimization using the initial guess from before
    opt_result_1 = minimize(
        objective_function,
        weights,
        method='SLSQP',
        bounds=bounds,
        constraints=constraints
    )

    # opt_result_2 = minimize(
    #     objective_function,
    #     weights,
    #     method='trust-constr',
    #     bounds=bounds,
    #     constraints=constraints
    # )

    opt_result = opt_result_1

    # if opt_result_1.fun < opt_result_2.fun:
    #     opt_result = opt_result_1
    # else:
    #     opt_result = opt_result_2

    new_weights_temp = list(opt_result.x)
    new_weights = []
    for i in new_weights_temp:
        new_weights.append(round(i,4))
    
    return new_weights, opt_result

# Min Variance

def min_variance(average, cov_matrix, num_symbs, expected_returns):

    def calculate_e3(weights, average):
        return np.dot(weights, average)
    
    def calculate_e4(weights, average):
        return ((1+calculate_e3(weights, average))**52)-1

    def calculate_e5(weights, cov_matrix):
        return np.dot(np.dot(weights, cov_matrix), weights)

    def objective_function(weights):
        e5 = calculate_e5(weights, cov_matrix)
        return e5

    weights = [1/num_symbs]*num_symbs

    # Constraints and bounds remain the same as before
    constraints = (
        {'type': 'eq', 'fun': lambda x: np.sum(x) - 1},
        {'type': 'ineq', 'fun': lambda x: calculate_e4(x, average) - expected_returns}
    )

    bounds = [(0, 1)] * len(weights)

    # Perform the optimization using the initial guess from before
    opt_result_1 = minimize(
        objective_function,
        weights,
        method='SLSQP',
        bounds=bounds,
        constraints=constraints
    )

    # opt_result_2 = minimize(
    #     objective_function,
    #     weights,
    #     method='trust-constr',
    #     bounds=bounds,
    #     constraints=constraints
    # )
    opt_result = opt_result_1

    # if opt_result_1.fun < opt_result_2.fun:
    #     opt_result = opt_result_1
    # else:
    #     opt_result = opt_result_2

    new_weights_temp = list(opt_result.x)
    new_weights = []
    for i in new_weights_temp:
        new_weights.append(round(i,4))
    
    return new_weights, opt_result

# Max Returns

def max_returns(average, cov_matrix, num_symbs, maximum_ann_stdev):

    def calculate_e3(weights, average):
        return np.dot(weights, average)

    def calculate_e5(weights, cov_matrix):
        return np.dot(np.dot(weights, cov_matrix), weights)
    
    def calculate_ann_stddev(weights, cov_matrix):
        e5 = calculate_e5(weights, cov_matrix)
        ann_stddev = math.sqrt(e5*52)
        return ann_stddev

    def objective_function(weights):
        e3 = calculate_e3(weights, average)
        return -e3

    weights = [1/num_symbs]*num_symbs

    # Constraints and bounds remain the same as before
    constraints = (
        {'type': 'eq', 'fun': lambda x: np.sum(x) - 1},
        {'type': 'ineq', 'fun': lambda x: maximum_ann_stdev - calculate_ann_stddev(x, cov_matrix)}
    )

    bounds = [(0, 1)] * len(weights)

    # Perform the optimization using the initial guess from before

    opt_result_1 = minimize(
        objective_function,
        weights,
        method='SLSQP',
        bounds=bounds,
        constraints=constraints
    )

    opt_result = opt_result_1

    # opt_result_2 = minimize(
    #     objective_function,
    #     weights,
    #     method='trust-constr',
    #     bounds=bounds,
    #     constraints=constraints
    # )

    # if opt_result_1.fun < opt_result_2.fun:
    #     opt_result = opt_result_1
    # else:
    #     opt_result = opt_result_2

    new_weights_temp = list(opt_result.x)
    new_weights = []
    for i in new_weights_temp:
        new_weights.append(round(i,4))
    
    return new_weights, opt_result

def delete_cells(sheet, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        cell_to_delete = f"{xl.utils.get_column_letter(col)}{row}"
        sheet[cell_to_delete].value = None

# main

def main(input_file_path, year):

    input_filename = os.path.basename(input_file_path)
    input_filedir = os.path.dirname(input_file_path)
    output_file_path = input_filedir+'/updated_'+input_filename
    add_sheets(year, input_file_path)

    # Getting end_char and num_symbs

    wb = xl.load_workbook(output_file_path)
    summary_stats_sheet = wb['summary_stats']
    first_row_values = [cell.value for cell in summary_stats_sheet[1]]
    li = [value for value in first_row_values if value is not None]
    wb.close()

    end_char = excel_column_name(len(li) + 1)
    num_symbs = len(li)

    # Getting average and cov_matrix

    wb = xw.Book(output_file_path)
    summary_stats_sheet = wb.sheets['summary_stats']
    average = summary_stats_sheet.range(f'B2:{end_char}2').value
    average = np.array(average)
    wb.close()

    cov_matrix = pd.read_excel(output_file_path, sheet_name='covariance_matrix', index_col=0)
    cov_matrix = cov_matrix.values

    risk_free_rate = nominal_risk_free_rate[year]
    max_sha_weights, max_sharpe_ratio_ = max_sharpe_ratio(average, cov_matrix, num_symbs, risk_free_rate)

    add_sheets(year, output_file_path, 'MAX SHA', max_sha_weights)

    expected_annual_returns = 0.1
    min_var_weights, min_variance_ = min_variance(average, cov_matrix, num_symbs, expected_annual_returns)
    add_sheets(year, input_filedir+'/updated_updated_'+input_filename, 'MIN VAR', max_sha_weights, min_var_weights, None, expected_annual_returns)

    max_annnual_std_dev = 0.4
    max_ret_weights, max_returns_ = max_returns(average, cov_matrix, num_symbs, max_annnual_std_dev)
    add_sheets(year, input_filedir+'/updated_updated_updated_'+input_filename, 'MAX RET', max_sha_weights, min_var_weights, max_ret_weights, expected_annual_returns, max_annnual_std_dev)
    
    os.remove(input_filedir+'/updated_'+input_filename)
    os.remove(input_filedir+'/updated_updated_'+input_filename)
    os.remove(input_filedir+'/updated_updated_updated_'+input_filename)
    os.rename(input_filedir+'/updated_updated_updated_updated_'+input_filename, input_filedir+'/'+input_filename[:-5]+'_optimized.xlsx')

    excel_file_path = input_filedir+'/'+input_filename[:-5]+'_optimized.xlsx'
    wb = xl.load_workbook(excel_file_path)

    opt_sheets = [wb['max_sharpe_ratio'], wb['min_variance'], wb['max_returns']]
    for sheet in opt_sheets:
        delete_cells(sheet, 1, 3, 5)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for col in sheet.columns:
            column = col[0].column_letter
            if sheet in opt_sheets:
                sheet.column_dimensions[column].width = 23.33
            else:
                sheet.column_dimensions[column].width = 20

    wb['covariance_matrix']['A1'].value = None

    wb.save(excel_file_path)

# get all files and call the main

def list_all_files(base_path):
    all_files = []
    for root, dirs, files in os.walk(base_path):
        for file in files:
            all_files.append(os.path.join(root, file))
    return all_files

def annual_to_daily_rate(value):
    return ((value+1)**(1/52))-1

nominal_risk_free_rate = {
    '2018': annual_to_daily_rate(0.0765),
    '2019': annual_to_daily_rate(0.0729),
    '2020': annual_to_daily_rate(0.0696),
    '2021': annual_to_daily_rate(0.0673),
    '2022': annual_to_daily_rate(0.0682),
    '2023': annual_to_daily_rate(0.0672)
}

path = '/Users/oomrawat/Desktop/04_FSP/FINC308 - Investment Analysis/Group Project/Industry Wise Data'

all_files_list = list_all_files(path)
all_files_w_year = []

for file in all_files_list:
    if file[-8:] != 'DS_Store':
        year = os.path.dirname(file)[-7:]
        if year[0] == '/':
            year = year[1:]
        all_files_w_year.append((file,year))

all_files_w_year = [(item,year) for item,year in all_files_w_year if not item.endswith('_optimized.xlsx')]

count = 1
for file in all_files_w_year:
    count += 1
    year = file[1][:4]
    file_path = file[0]

    if not os.path.exists(file_path[:-5]+'_optimized.xlsx'):

        print(count, file[1])
        
        source_path = file_path
        destination_path = os.path.dirname(file_path)+'/temp.xlsx'

        shutil.copy(source_path, destination_path)

        main(destination_path, year)

        new_file_name = source_path[:-5] + '_optimized.xlsx'
        os.rename(destination_path[:-5] + '_optimized.xlsx', new_file_name)